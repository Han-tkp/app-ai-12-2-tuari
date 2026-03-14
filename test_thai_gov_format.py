"""
Test Excel Export - Thai Government Format
Based on: แบบบันทึกประเมินเครื่องปี 2569
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import numpy as np
from datetime import datetime
import random

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "ผลการวิเคราะห์ละอองฝอย"

# Thai headers (from the reference image)
HEADERS = {
    'title': 'แบบบันทึกผลการวิเคราะห์ละอองฝอยสารเคมี',
    'subtitle': 'ตามมาตรฐาน WHO Chemical Spray Droplet Analysis',
    'project': 'โครงการทดสอบ',
    'date': f'วันที่: {datetime.now().strftime("%d/%m/%Y")}',
    
    # Column headers
    'col1': 'ละอองที่',
    'col2': 'เส้นผ่านศูนย์กลางที่สำรวจได้\n(2 เท่าของรัศมี)',
    'col3': 'เส้นผ่านศูนย์กลางจริง\n(หารด้วย 1.15)\n(อ่านค่านี้)',
    'col4': 'ปริมาตรที่คำนวณ\n(ปริมาตรทรงกลม = 4/3πR³)',
    'col5': 'ผลรวมสะสมของปริมาตรในแต่ละชั้น\n(ปริมาตรสะสม)',
    'col6': '% ของปริมาตรในแต่ละชั้น\n(เพื่อหาจุดที่เป็น 50%\nของปริมาตรสะสม)',
}

# Styles
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
red_font = Font(color="FF0000", bold=True, size=11)
black_font = Font(size=11)
border_thin = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Generate test data: 300 droplets
random.seed(42)  # For reproducibility
np.random.seed(42)

# Generate sizes with normal distribution around 20-25 µm
sizes = np.random.normal(22, 4, 300)
sizes = np.clip(sizes, 10.0, 31.0)

droplets = []
for size in sizes:
    size = size + random.uniform(-0.3, 0.3)
    size = max(10.0, min(31.0, size))
    droplets.append(round(size, 4))

# Sort by size (small to large)
droplets.sort()

# ── Title Section ─────────────────────────────────────────────────────────────
ws.merge_cells("A1:F1")
ws["A1"] = HEADERS['title']
ws["A1"].font = Font(bold=True, size=16)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:F2")
ws["A2"] = HEADERS['subtitle']
ws["A2"].font = Font(italic=True, size=12)
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 25

ws.merge_cells("A3:F3")
ws["A3"] = f"{HEADERS['project']}  |  {HEADERS['date']}"
ws["A3"].font = Font(size=11)
ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[3].height = 20

# ── Column Headers (Row 4-7) ──────────────────────────────────────────────────
# Row 4: Main headers
ws.merge_cells("A4:A7")
ws["A4"] = HEADERS['col1']
ws["A4"].font = black_font
ws["A4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws["A4"].border = border_thin

ws.merge_cells("B4:B7")
ws["B4"] = HEADERS['col2']
ws["B4"].font = red_font
ws["B4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws["B4"].border = border_thin
ws["B4"].fill = header_fill

ws.merge_cells("C4:C7")
ws["C4"] = HEADERS['col3']
ws["C4"].font = red_font
ws["C4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws["C4"].border = border_thin
ws["C4"].fill = header_fill

ws.merge_cells("D4:D7")
ws["D4"] = HEADERS['col4']
ws["D4"].font = red_font
ws["D4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws["D4"].border = border_thin
ws["D4"].fill = header_fill

ws.merge_cells("E4:E7")
ws["E4"] = HEADERS['col5']
ws["E4"].font = red_font
ws["E4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws["E4"].border = border_thin
ws["E4"].fill = header_fill

ws.merge_cells("F4:F7")
ws["F4"] = HEADERS['col6']
ws["F4"].font = red_font
ws["F4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws["F4"].border = border_thin
ws["F4"].fill = header_fill

# Row 5: Column numbers (1), (2), (3), etc.
ws["A8"] = "(1)"
ws["B8"] = "(2)"
ws["C8"] = "(3)"
ws["D8"] = "(4)"
ws["E8"] = "(5)"
ws["F8"] = "(6)"

for col in range(1, 7):
    cell = ws.cell(row=8, column=col)
    cell.font = Font(italic=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_thin

ws.row_dimensions[8].height = 25

# ── Data Rows ─────────────────────────────────────────────────────────────────
# Calculate volumes and cumulative
volumes = []
for d in droplets:
    radius = d / 2  # Radius = diameter / 2
    volume = (4/3) * np.pi * (radius ** 3)
    volumes.append(volume)

# Cumulative volume
cumulative_volumes = np.cumsum(volumes)
total_volume = cumulative_volumes[-1]

# Cumulative percentage
cumulative_pct = (cumulative_volumes / total_volume) * 100

# True diameter (divide by 1.15 - Spread Factor correction)
true_diameters = [d / 1.15 for d in droplets]

# Write data (show first 100 rows for demo)
start_row = 9
display_count = min(100, len(droplets))

for i in range(display_count):
    row = start_row + i
    
    # Column 1: Droplet number
    ws.cell(row=row, column=1, value=i+1)
    
    # Column 2: Measured diameter
    ws.cell(row=row, column=2, value=droplets[i])
    
    # Column 3: True diameter (corrected)
    ws.cell(row=row, column=3, value=true_diameters[i])
    
    # Column 4: Volume
    ws.cell(row=row, column=4, value=volumes[i])
    
    # Column 5: Cumulative volume
    ws.cell(row=row, column=5, value=cumulative_volumes[i])
    
    # Column 6: Cumulative percentage
    ws.cell(row=row, column=6, value=cumulative_pct[i])
    
    # Apply formatting
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
        
        # Number formatting
        if col == 1:
            cell.number_format = '0'
        elif col in [2, 3]:
            cell.number_format = '0.0000'
        elif col == 4:
            cell.number_format = '0.000'
        elif col == 5:
            cell.number_format = '0.000'
        elif col == 6:
            cell.number_format = '0.00'

    ws.row_dimensions[row].height = 20

# Column widths
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 20

# ── Summary Statistics (below data) ───────────────────────────────────────────
summary_row = start_row + display_count + 2

ws.merge_cells(f"A{summary_row}:F{summary_row}")
ws[f"A{summary_row}"] = "สรุปผลการวิเคราะห์"
ws[f"A{summary_row}"].font = Font(bold=True, size=14)
ws[f"A{summary_row}"].alignment = Alignment(horizontal="left")
ws.row_dimensions[summary_row].height = 25

# Statistics
stats = [
    ("จำนวนละอองทั้งหมด", len(droplets), "จุด"),
    ("ขนาดเฉลี่ย (Mean)", np.mean(droplets), "µm"),
    ("VMD (Dv0.5)", np.percentile(droplets, 50), "µm"),
    ("Dv0.1", np.percentile(droplets, 10), "µm"),
    ("Dv0.9", np.percentile(droplets, 90), "µm"),
    ("SPAN", (np.percentile(droplets, 90) - np.percentile(droplets, 10)) / np.percentile(droplets, 50), ""),
    ("จำนวนในช่วง WHO (10-30 µm)", sum(1 for d in droplets if 10 <= d <= 30), f"จุด ({sum(1 for d in droplets if 10 <= d <= 30)/len(droplets)*100:.1f}%)"),
]

for idx, (label, value, unit) in enumerate(stats):
    row = summary_row + 1 + idx
    ws[f"A{row}"] = label
    ws[f"A{row}"].font = Font(bold=True, size=11)
    ws[f"B{row}"] = f"{value:.4f}" if isinstance(value, float) else value
    ws[f"B{row}"].font = Font(size=11)
    if unit:
        ws[f"C{row}"] = unit
        ws[f"C{row}"].font = Font(size=11)
    ws.row_dimensions[row].height = 20

# ── Add Chart ─────────────────────────────────────────────────────────────────
from openpyxl.chart import BarChart, Reference

chart_row = summary_row + len(stats) + 2
ws.merge_cells(f"A{chart_row}:F{chart_row}")
ws[f"A{chart_row}"] = "การกระจายขนาดละออง"
ws[f"A{chart_row}"].font = Font(bold=True, size=12)
ws.row_dimensions[chart_row].height = 25

# Calculate distribution bins
bins = {"<10": 0, "10-15": 0, "15-20": 0, "20-25": 0, "25-30": 0, ">30": 0}
for d in droplets:
    if d < 10: bins["<10"] += 1
    elif d < 15: bins["10-15"] += 1
    elif d < 20: bins["15-20"] += 1
    elif d < 25: bins["20-25"] += 1
    elif d < 30: bins["25-30"] += 1
    else: bins[">30"] += 1

# Add bin data
bin_start_row = chart_row + 2
for i, (label, count) in enumerate(bins.items()):
    ws.cell(row=bin_start_row + i, column=8, value=label)
    ws.cell(row=bin_start_row + i, column=9, value=count)

# Create chart
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "การกระจายขนาดละอองฝอย"
chart.y_axis.title = "จำนวน (จุด)"
chart.x_axis.title = "ช่วงขนาด (µm)"

data_ref = Reference(ws, min_col=9, min_row=bin_start_row, max_row=bin_start_row+len(bins)-1, max_col=9)
cat_ref = Reference(ws, min_col=8, min_row=bin_start_row+1, max_row=bin_start_row+len(bins)-1)

chart.add_data(data_ref, titles_from_data=False)
chart.set_categories(cat_ref)
chart.shape = 4
chart.width = 14
chart.height = 10

ws.add_chart(chart, f"A{chart_row+1}")

# Save
output_path = "C:/Users/h4n/Desktop/webappsdesktop/datafordevapp/excel/Test_Thai_Government_Format.xlsx"
wb.save(output_path)

print(f"✓ Excel file created: {output_path}")
print(f"\nStructure:")
print(f"  Title: {HEADERS['title']}")
print(f"  Columns:")
print(f"    (1) ละอองที่")
print(f"    (2) เส้นผ่านศูนย์กลางที่สำรวจได้")
print(f"    (3) เส้นผ่านศูนย์กลางจริง (หารด้วย 1.15)")
print(f"    (4) ปริมาตรที่คำนวณ (4/3πR³)")
print(f"    (5) ผลรวมสะสมของปริมาตร")
print(f"    (6) % ของปริมาตรสะสม")
print(f"\nStats:")
print(f"  Total droplets: {len(droplets)}")
print(f"  Mean size: {np.mean(droplets):.4f} µm")
print(f"  VMD (Dv0.5): {np.percentile(droplets, 50):.4f} µm")
print(f"  Dv0.1: {np.percentile(droplets, 10):.4f} µm")
print(f"  Dv0.9: {np.percentile(droplets, 90):.4f} µm")
print(f"  SPAN: {(np.percentile(droplets, 90) - np.percentile(droplets, 10)) / np.percentile(droplets, 50):.4f}")
print(f"  In WHO range (10-30 µm): {sum(1 for d in droplets if 10 <= d <= 30)} ({sum(1 for d in droplets if 10 <= d <= 30)/len(droplets)*100:.1f}%)")
