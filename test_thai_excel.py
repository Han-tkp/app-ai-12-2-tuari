"""
Test Excel Export - Thai Language with 3 slides x 300 droplets
Including both AI and Manual detections
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import numpy as np
from datetime import datetime
import random

# Create workbook
wb = Workbook()

# Generate test data: 3 slides x 300 droplets each
# Mix of AI and Manual detections
# Size range: 10.00 - 31.00 µm (WHO range: 10-30)

def generate_droplet_data(count=300, manual_ratio=0.15):
    """Generate droplet data with mix of AI and Manual."""
    droplets = []
    
    # Generate sizes with normal distribution around 20-25 µm
    sizes = np.random.normal(22, 4, count)
    sizes = np.clip(sizes, 10.0, 31.0)  # Clamp to 10-31 range
    
    for i, size in enumerate(sizes):
        # Add some variation
        size = size + random.uniform(-0.5, 0.5)
        size = max(10.0, min(31.0, size))
        
        # 15% manual, 85% AI
        is_manual = random.random() < manual_ratio
        source = "Manual" if is_manual else "AI"
        
        droplets.append({
            'size': round(size, 3),
            'source': source
        })
    
    return droplets

# Generate data for 3 slides
slides_data = [
    {"name": "Slide 1", "droplets": generate_droplet_data(300, 0.15)},
    {"name": "Slide 2", "droplets": generate_droplet_data(300, 0.12)},
    {"name": "Slide 3", "droplets": generate_droplet_data(300, 0.18)},
]

# Thai labels
TH = {
    'title': 'รายงานการวิเคราะห์ละอองฝอยสารเคมี',
    'target': 'ขนาดเป้าหมาย',
    'total_slides': 'จำนวนสไลด์',
    'generated': 'สร้างเมื่อ',
    'slide_summary': 'สรุปผลสไลด์',
    'slide': 'สไลด์',
    'dv01': 'Dv0.1 (µm)',
    'dv05': 'Dv0.5 / VMD (µm)',
    'dv09': 'Dv0.9 (µm)',
    'span': 'SPAN',
    'count': 'จำนวน (n)',
    'who_status': 'สถานะ WHO',
    'pass': 'ผ่าน',
    'fail': 'ไม่ผ่าน',
    'in_range_pct': 'ในช่วง (%)',
    'overall_stats': 'สถิติรวม',
    'total_droplets': 'ละอองทั้งหมด',
    'avg_vmd': 'VMD เฉลี่ย (µm)',
    'avg_span': 'SPAN เฉลี่ย',
    'who_compliant': 'สไลด์ที่ผ่าน WHO',
    'vmd_chart_title': 'เปรียบเทียบ VMD แต่ละสไลด์',
    'vmd_chart_y': 'VMD (µm)',
    'vmd_chart_x': 'สไลด์',
    'droplet_data': 'ข้อมูลละออง (เรียงตามขนาด)',
    'who_status_short': 'สถานะ WHO',
    'in_range': 'ในช่วง',
    'hash': '#',
    'diameter': 'ขนาด (µm)',
    'who_range': 'ช่วง WHO',
    'volume': 'ปริมาตร (µm³)',
    'spread_factor': 'Spread Factor',
    'cumulative_pct': 'สะสม (%)',
    'yes': 'ใช่',
    'no': 'ไม่',
    'size_distribution': 'การกระจายขนาดละออง',
    'size_range': 'ช่วงขนาด (µm)',
    'count_label': 'จำนวน',
    'ai_count': 'AI',
    'manual_count': 'Manual',
}

def compute_stats(droplets):
    """Compute Dv10, Dv50, Dv90, SPAN for a list of droplet sizes."""
    sizes = [d['size'] for d in droplets]
    if not sizes:
        return {"dv10": None, "dv50": None, "dv90": None, "span": None, "count": 0, "ai": 0, "manual": 0}
    
    diams = sorted(sizes)
    vols = [(np.pi / 6) * (d ** 3) for d in diams]
    total_vol = sum(vols)
    cum = np.cumsum(vols) / total_vol
    
    dv10 = float(diams[int(np.searchsorted(cum, 0.1))])
    dv50 = float(diams[int(np.searchsorted(cum, 0.5))])
    dv90 = float(diams[int(np.searchsorted(cum, 0.9))])
    span = float((dv90 - dv10) / dv50) if dv50 > 0 else 0.0
    
    ai_count = sum(1 for d in droplets if d['source'] == 'AI')
    manual_count = sum(1 for d in droplets if d['source'] == 'Manual')
    
    return {
        "dv10": dv10, "dv50": dv50, "dv90": dv90, "span": span,
        "count": len(diams), "ai": ai_count, "manual": manual_count
    }

stats_list = [compute_stats(s["droplets"]) for s in slides_data]

# ── Sheet 1: Summary Dashboard ───────────────────────────────────────────────
ws = wb.active
ws.title = "Summary"

# Title
ws.merge_cells("A1:F1")
ws["A1"] = "DropDetect AI - ทดสอบ"
ws["A1"].font = Font(bold=True, size=16)
ws.row_dimensions[1].height = 30

# Subtitle
ws.merge_cells("A2:F2")
ws["A2"] = TH['title']
ws["A2"].font = Font(italic=True, size=12)
ws.row_dimensions[2].height = 20

# Info
ws.merge_cells("A3:F3")
ws["A3"] = f"{TH['target']}: 224 µm  |  {TH['total_slides']}: {len(slides_data)}  |  {TH['generated']}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
ws.row_dimensions[3].height = 18

# SLIDE SUMMARY section
ws.merge_cells("A5:F5")
ws["A5"] = TH['slide_summary']
ws["A5"].font = Font(bold=True, size=12)
ws.row_dimensions[5].height = 24

# Header
header_row = 6
headers = [TH['slide'], TH['dv01'], TH['dv05'], TH['dv09'], TH['span'], TH['count'], TH['who_status'], TH['in_range_pct']]
for col_idx, h in enumerate(headers, 1):
    c = ws.cell(row=header_row, column=col_idx, value=h)
    c.font = Font(bold=True, size=10)
    c.alignment = Alignment(horizontal="center")
    c.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
ws.row_dimensions[header_row].height = 22

# Data rows
for row_idx, (slide, st) in enumerate(zip(slides_data, stats_list), header_row + 1):
    who_pass = st["dv50"] is not None and 10 <= st["dv50"] <= 30
    who_text = TH['pass'] if who_pass else TH['fail']
    in_range = sum(1 for d in slide["droplets"] if 10 <= d["size"] <= 30)
    in_range_pct = (in_range / len(slide["droplets"]) * 100) if slide["droplets"] else 0
    
    ws.cell(row=row_idx, column=1, value=slide["name"]).font = Font(bold=True, size=10)
    ws.cell(row=row_idx, column=2, value=round(st['dv10'], 3))
    ws.cell(row=row_idx, column=3, value=round(st['dv50'], 3))
    ws.cell(row=row_idx, column=4, value=round(st['dv90'], 3))
    ws.cell(row=row_idx, column=5, value=round(st['span'], 4))
    ws.cell(row=row_idx, column=6, value=st['count'])
    
    who_cell = ws.cell(row=row_idx, column=7, value=who_text)
    who_cell.font = Font(bold=True, color="008000" if who_pass else "FF0000")
    
    pct_cell = ws.cell(row=row_idx, column=8, value=round(in_range_pct, 1))
    pct_cell.number_format = "0.0%"
    
    for col in range(1, 9):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    ws.row_dimensions[row_idx].height = 20

# Column widths
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 10
ws.column_dimensions["G"].width = 14
ws.column_dimensions["H"].width = 12

# OVERALL STATISTICS
overall_row = header_row + len(slides_data) + 2
total_droplets = sum(st['count'] for st in stats_list)
total_ai = sum(st['ai'] for st in stats_list)
total_manual = sum(st['manual'] for st in stats_list)
avg_vmd = sum(st['dv50'] for st in stats_list) / len(stats_list)
avg_span = sum(st['span'] for st in stats_list) / len(stats_list)
who_pass_count = sum(1 for st in stats_list if st['dv50'] and 10 <= st['dv50'] <= 30)

ws.merge_cells(f"A{overall_row}:F{overall_row}")
ws[f"A{overall_row}"] = TH['overall_stats']
ws[f"A{overall_row}"].font = Font(bold=True, size=12)
ws.row_dimensions[overall_row].height = 24

overall_data = [
    (TH['total_droplets'], total_droplets),
    (f"  - {TH['ai_count']}", total_ai),
    (f"  - {TH['manual_count']}", total_manual),
    (TH['avg_vmd'], round(avg_vmd, 3)),
    (TH['avg_span'], round(avg_span, 4)),
    (TH['who_compliant'], f"{who_pass_count}/{len(slides_data)}"),
]

for col_idx, (label, value) in enumerate(overall_data, 1):
    row = overall_row + 1 + (col_idx - 1) // 2
    col = ((col_idx - 1) % 2) * 2 + 1
    ws.cell(row=row, column=col, value=label).font = Font(bold=True, size=10)
    ws.cell(row=row, column=col+1, value=value).font = Font(size=10)

# Add VMD Distribution Chart
chart_row_start = overall_row + 5
ws.merge_cells(f"A{chart_row_start}:F{chart_row_start}")
ws[f"A{chart_row_start}"] = TH['vmd_chart_title']
ws[f"A{chart_row_start}"].font = Font(bold=True, size=12)
ws.row_dimensions[chart_row_start].height = 24

chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = TH['vmd_chart_title']
chart.y_axis.title = TH['vmd_chart_y']
chart.x_axis.title = TH['vmd_chart_x']

data_ref = Reference(ws, min_col=3, min_row=header_row+1, max_row=header_row+len(slides_data), max_col=3)
cat_ref = Reference(ws, min_col=1, min_row=header_row+1, max_row=header_row+len(slides_data))

chart.add_data(data_ref, titles_from_data=False)
chart.set_categories(cat_ref)
chart.shape = 4
chart.width = 14
chart.height = 10

ws.add_chart(chart, f"A{chart_row_start+1}")

# ── Per-slide sheets ─────────────────────────────────────────────────────────
for slide_idx, slide in enumerate(slides_data):
    # Sort by size (small to large)
    sorted_droplets = sorted(slide["droplets"], key=lambda x: x["size"])
    st = stats_list[slide_idx]
    sheet_name = slide["name"][:31]
    ws2 = wb.create_sheet(title=sheet_name)
    
    # Title
    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"{slide['name']} - {TH['droplet_data']}"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.row_dimensions[1].height = 28
    
    # Stats summary
    stat_labels = [
        (TH['dv01'], round(st['dv10'], 3)),
        (TH['dv05'], round(st['dv50'], 3)),
        (TH['dv09'], round(st['dv90'], 3)),
        (TH['span'], round(st['span'], 4)),
        (TH['count'], st["count"]),
        (f"  - {TH['ai_count']}", st['ai']),
        (f"  - {TH['manual_count']}", st['manual']),
    ]
    for r, (lbl, val) in enumerate(stat_labels, 2):
        ws2.cell(row=r, column=1, value=lbl).font = Font(bold=True, size=10)
        ws2.cell(row=r, column=2, value=val).font = Font(size=10)
        ws2.row_dimensions[r].height = 18
    
    # WHO status
    who_pass = st['dv50'] and 10 <= st['dv50'] <= 30
    in_range = sum(1 for d in sorted_droplets if 10 <= d["size"] <= 30)
    in_range_pct = (in_range / len(sorted_droplets) * 100)
    
    ws2.cell(row=2, column=3, value=TH['who_status_short']).font = Font(bold=True, size=10)
    ws2.cell(row=2, column=4, value=TH['pass'] if who_pass else TH['fail']).font = Font(bold=True, color="008000" if who_pass else "FF0000")
    
    ws2.cell(row=3, column=3, value=TH['in_range']).font = Font(bold=True, size=10)
    ws2.cell(row=3, column=4, value=round(in_range_pct, 1)).number_format = "0.0%"
    
    # Raw data header
    hdr_row = 10
    headers = [TH['hash'], TH['diameter'], TH['who_range'], TH['volume'], TH['spread_factor'], TH['cumulative_pct'], 'Source']
    for ci, txt in enumerate(headers, 1):
        c = ws2.cell(row=hdr_row, column=ci, value=txt)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    ws2.row_dimensions[hdr_row].height = 20
    
    # Data rows (first 100 for demo - sorted small to large)
    cum_count = 0
    display_count = min(100, len(sorted_droplets))  # Show first 100 for demo
    
    for i, d in enumerate(sorted_droplets[:display_count], 1):
        cum_count += 1
        size = d["size"]
        vol = (np.pi / 6) * (size ** 3)
        sf = 0.86 if size > 20 else 0.80 if size >= 15 else 0.75 if size >= 10 else 0.70
        in_range_flag = 10.0 <= size <= 30.0
        cum_pct = (cum_count / len(sorted_droplets) * 100)
        row_num = hdr_row + i
        
        ws2.cell(row=row_num, column=1, value=i)
        ws2.cell(row=row_num, column=2, value=size).number_format = "0.000"
        ws2.cell(row=row_num, column=3, value=TH['yes'] if in_range_flag else TH['no']).font = Font(color="008000" if in_range_flag else "FF0000")
        ws2.cell(row=row_num, column=4, value=round(vol, 1)).number_format = "0.0"
        ws2.cell(row=row_num, column=5, value=sf).number_format = "0.00"
        ws2.cell(row=row_num, column=6, value=round(cum_pct, 1)).number_format = "0.0%"
        ws2.cell(row=row_num, column=7, value=d["source"]).font = Font(color="0000FF" if d["source"] == "AI" else "FF00FF")
        
        for ci in range(1, 8):
            cell = ws2.cell(row=row_num, column=ci)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws2.row_dimensions[row_num].height = 16
    
    # Column widths
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 12
    ws2.column_dimensions["G"].width = 10
    
    # Add distribution chart
    bins = {"<10": 0, "10-15": 0, "15-20": 0, "20-25": 0, "25-30": 0, ">30": 0}
    for d in sorted_droplets:
        size = d["size"]
        if size < 10: bins["<10"] += 1
        elif size < 15: bins["10-15"] += 1
        elif size < 20: bins["15-20"] += 1
        elif size < 25: bins["20-25"] += 1
        elif size < 30: bins["25-30"] += 1
        else: bins[">30"] += 1
    
    start_col = 8
    start_row = hdr_row + display_count + 2
    ws2.cell(row=start_row, column=start_col, value=TH['size_range']).font = Font(bold=True)
    ws2.cell(row=start_row, column=start_col+1, value=TH['count_label']).font = Font(bold=True)
    
    for i, (label, count) in enumerate(zip(list(bins.keys()), list(bins.values())), start_row+1):
        ws2.cell(row=i, column=start_col, value=label)
        ws2.cell(row=i, column=start_col+1, value=count)
    
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = TH['size_distribution']
    chart2.y_axis.title = TH['count_label']
    chart2.x_axis.title = TH['size_range']
    
    data_ref2 = Reference(ws2, min_col=start_col+1, min_row=start_row, max_row=start_row+len(bins), max_col=start_col+1)
    cat_ref2 = Reference(ws2, min_col=start_col, min_row=start_row+1, max_row=start_row+len(bins))
    
    chart2.add_data(data_ref2, titles_from_data=False)
    chart2.set_categories(cat_ref2)
    chart2.shape = 4
    chart2.width = 12
    chart2.height = 8
    
    ws2.add_chart(chart2, f"A{start_row+1}")

# Save
output_path = "C:/Users/h4n/Desktop/webappsdesktop/datafordevapp/excel/Test_Thai_3Slides_300Droplets.xlsx"
wb.save(output_path)
print(f"✓ Excel file created: {output_path}")
print(f"\nStructure:")
print(f"  Sheet 1: Summary (Dashboard + Chart)")
print(f"  Sheet 2: Slide 1 ({len(slides_data[0]['droplets'])} droplets)")
print(f"  Sheet 3: Slide 2 ({len(slides_data[1]['droplets'])} droplets)")
print(f"  Sheet 4: Slide 3 ({len(slides_data[2]['droplets'])} droplets)")
print(f"\nStats:")
for i, (slide, st) in enumerate(zip(slides_data, stats_list), 1):
    in_range = sum(1 for d in slide["droplets"] if 10 <= d["size"] <= 30)
    in_range_pct = (in_range / len(slide["droplets"]) * 100)
    print(f"  Slide {i}: VMD={st['dv50']:.3f}, SPAN={st['span']:.4f}, Count={st['count']}, In Range={in_range_pct:.1f}%")
    print(f"    - AI: {st['ai']}, Manual: {st['manual']}")
