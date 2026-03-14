import cv2
import asyncio
import base64
import numpy as np
import onnxruntime as ort
import supervision as sv
import psutil
import json
import os
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
import logging
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

logger = logging.getLogger("dropdetect")

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Single-threaded executor: ONNX inference runs here, off the event loop
_inference_executor = ThreadPoolExecutor(max_workers=1)

# Profile parameters — resolution stays 640×640 on every profile
PROFILE_SKIP = {"low": 3, "mid": 2, "high": 1}
PROFILE_QUEUE_SIZE = {"low": 2, "mid": 4, "high": 8}
# track_buffer scaled so all profiles retain ~3 s of track history at 30 FPS camera
PROFILE_TRACK_BUFFER = {"low": 90, "mid": 60, "high": 30}

# Letterbox constant: 640×480 camera → 640×640 model input (80px top+bottom padding)
PAD_TOP = 80

# ── Language Settings ────────────────────────────────────────────────────────
class LanguageConfig:
    """Configuration for Excel export language."""
    def __init__(self, lang='th'):
        self.lang = lang
        
        # Thai labels
        self.TH = {
            'title': 'รายงานการวิเคราะห์ละอองฝอยสารเคมี',
            'target': 'ขนาดเป้าหมาย',
            'total_slides': 'จำนวนสไลด์',
            'generated': 'สร้างเมื่อ',
            'slide_summary': 'สรุปผลสไลด์',
            'slide': 'สไลด์',
            'dv01': 'Dv0.1 (µm)',
            'dv05': 'Dv0.5 / VMD (µm)',
            'dv09': 'Dv0.9 (µm)',
            'span': 'SPAN',
            'count': 'จำนวน (n)',
            'who_status': 'สถานะ WHO',
            'pass': 'ผ่าน',
            'fail': 'ไม่ผ่าน',
            'in_range_pct': 'ในช่วง (%)',
            'overall_stats': 'สถิติรวม',
            'total_droplets': 'ละอองทั้งหมด',
            'avg_vmd': 'VMD เฉลี่ย (µm)',
            'avg_span': 'SPAN เฉลี่ย',
            'who_compliant': 'สไลด์ที่ผ่าน WHO',
            'vmd_chart_title': 'เปรียบเทียบ VMD แต่ละสไลด์',
            'vmd_chart_y': 'VMD (µm)',
            'vmd_chart_x': 'สไลด์',
            # Per-slide sheet
            'droplet_data': 'ข้อมูลละออง (เรียงตามขนาด)',
            'who_status_short': 'สถานะ WHO',
            'in_range': 'ในช่วง',
            'hash': '#',
            'diameter': 'ขนาด (µm)',
            'who_range': 'ช่วง WHO',
            'volume': 'ปริมาตร (µm³)',
            'spread_factor': 'Spread Factor',
            'cumulative_pct': 'สะสม (%)',
            'yes': 'ใช่',
            'no': 'ไม่',
            'size_distribution': 'การกระจายขนาดละออง',
            'size_range': 'ช่วงขนาด (µm)',
            'count_label': 'จำนวน',
            'ai_count': 'AI',
            'manual_count': 'Manual',
        }
        
        # English labels
        self.EN = {
            'title': 'WHO Chemical Spray Droplet Analysis Report',
            'target': 'Target Size',
            'total_slides': 'Total Slides',
            'generated': 'Generated',
            'slide_summary': 'SLIDE SUMMARY',
            'slide': 'Slide',
            'dv01': 'Dv0.1 (µm)',
            'dv05': 'Dv0.5 / VMD (µm)',
            'dv09': 'Dv0.9 (µm)',
            'span': 'SPAN',
            'count': 'Count (n)',
            'who_status': 'WHO Status',
            'pass': 'Pass',
            'fail': 'Fail',
            'in_range_pct': 'In Range %',
            'overall_stats': 'OVERALL STATISTICS',
            'total_droplets': 'Total Droplets',
            'avg_vmd': 'Avg VMD (µm)',
            'avg_span': 'Avg SPAN',
            'who_compliant': 'WHO Compliant Slides',
            'vmd_chart_title': 'VMD Comparison by Slide',
            'vmd_chart_y': 'VMD (µm)',
            'vmd_chart_x': 'Slide',
            # Per-slide sheet
            'droplet_data': 'Droplet Data (Sorted by Size)',
            'who_status_short': 'WHO Status',
            'in_range': 'In Range',
            'hash': '#',
            'diameter': 'Diameter (µm)',
            'who_range': 'WHO Range',
            'volume': 'Volume (µm³)',
            'spread_factor': 'Spread Factor',
            'cumulative_pct': 'Cumulative %',
            'yes': 'Yes',
            'no': 'No',
            'size_distribution': 'Droplet Size Distribution',
            'size_range': 'Size Range (µm)',
            'count_label': 'Count',
            'ai_count': 'AI',
            'manual_count': 'Manual',
        }
    
    def get(self, key):
        """Get label in current language."""
        labels = self.TH if self.lang == 'th' else self.EN
        return labels.get(key, key)

# Default language config
DEFAULT_LANG = LanguageConfig('th')


def open_camera(idx: int) -> cv2.VideoCapture:
    cap = cv2.VideoCapture(idx, cv2.CAP_DSHOW)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    cap.set(cv2.CAP_PROP_FPS, 30)
    actual_w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    actual_h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    actual_fps = cap.get(cv2.CAP_PROP_FPS)
    logger.info("Camera %d opened: %dx%d @ %.0f FPS", idx, actual_w, actual_h, actual_fps)
    return cap


def detect_profile() -> str:
    ram_gb = psutil.virtual_memory().total / (1024 ** 3)
    if ram_gb < 6:
        return "low"
    elif ram_gb < 12:
        return "mid"
    else:
        return "high"


# --- Pydantic Models ---
class SlideData(BaseModel):
    id: str
    name: str
    droplets: list[float]

class SaveProjectRequest(BaseModel):
    project_name: str
    target_size: int
    save_directory: str
    slides: list[SlideData]
    language: str = "th"  # 'th' or 'en'

class ManualDataRequest(BaseModel):
    data: list[float] = []

class DropletSystem:
    def __init__(self):
        self.lens = "10x"
        self.session = None
        self.calibration_value = 2.7926330340561596e-07
        self.conf_threshold = 0.25
        self.profile = detect_profile()
        self.tracker = sv.ByteTrack(lost_track_buffer=PROFILE_TRACK_BUFFER[self.profile])
        self.counted_ids = set()
        self.droplet_data = {}
        self.pending_snapshot = False
        self.is_ai_active = False
        self.inference_skip = PROFILE_SKIP[self.profile]
        self.frame_counter = 0
        self.load_resources("10x")
        logger.info("DropletSystem init — profile=%s  RAM=%.1f GB  skip=1/%d",
                    self.profile,
                    psutil.virtual_memory().total / (1024 ** 3),
                    self.inference_skip)

    def set_profile(self, profile: str):
        if profile in PROFILE_SKIP:
            self.profile = profile
            self.inference_skip = PROFILE_SKIP[profile]
            self.tracker = sv.ByteTrack(lost_track_buffer=PROFILE_TRACK_BUFFER[profile])
            logger.info("Profile overridden to %s (skip=1/%d, track_buffer=%d)", profile, self.inference_skip, PROFILE_TRACK_BUFFER[profile])

    def load_resources(self, lens: str):
        self.lens = lens
        model_path = os.path.join(BASE_DIR, "fileonnx", f"yolov8n_{lens}.onnx")
        json_path = os.path.join(BASE_DIR, "resources", f"{lens}.json")
        if os.path.exists(model_path):
            self.session = ort.InferenceSession(model_path, providers=['CPUExecutionProvider'])
            self.input_name = self.session.get_inputs()[0].name
        if os.path.exists(json_path):
            with open(json_path, 'r') as f:
                data = json.load(f)
                for cal in data['calibrations']:
                    if cal['name'].lower() == lens.lower():
                        self.calibration_value = cal['value']
        self.reset_stats()

    def reset_stats(self):
        self.tracker = sv.ByteTrack(lost_track_buffer=PROFILE_TRACK_BUFFER[self.profile])
        self.counted_ids = set()
        self.droplet_data = {}

    def calculate_stats(self):
        if not self.droplet_data: return 0.0, 0.0, 0, [0,0,0], 0
        diams = sorted(self.droplet_data.values())
        vols = [(np.pi / 6) * (float(d)**3) for d in diams]
        total_vol = sum(vols)
        out_of_bounds = sum(1 for d in diams if d < 10.0 or d > 30.0)
        out_of_bounds_pct = (out_of_bounds / len(diams)) * 100 if len(diams) > 0 else 0
        if total_vol == 0: return 0.0, 0.0, len(diams), [0,0,0], round(out_of_bounds_pct, 1)
        cum_vol = np.cumsum(vols) / total_vol
        dv10 = float(diams[np.searchsorted(cum_vol, 0.1)])
        dv50 = float(diams[np.searchsorted(cum_vol, 0.5)])
        dv90 = float(diams[np.searchsorted(cum_vol, 0.9)])
        span = float((dv90 - dv10) / dv50) if dv50 > 0 else 0.0
        return dv50, span, len(diams), [dv10, dv50, dv90], round(out_of_bounds_pct, 1)

    def update_manual_data(self, manual_list):
        # Negative IDs represent manual annotations to avoid collision with tracker IDs
        self.droplet_data = {k: v for k, v in self.droplet_data.items() if k >= 0}
        for i, d in enumerate(manual_list):
            self.droplet_data[-(i+1000)] = float(d)

system = DropletSystem()

def _compute_slide_stats(droplets: list[float]) -> dict:
    """Compute Dv10, Dv50, Dv90, SPAN for a list of diameters."""
    if not droplets:
        return {"dv10": None, "dv50": None, "dv90": None, "span": None, "count": 0}
    diams = sorted(droplets)
    vols = [(np.pi / 6) * (d ** 3) for d in diams]
    total_vol = sum(vols)
    cum = np.cumsum(vols) / total_vol
    dv10 = float(diams[int(np.searchsorted(cum, 0.1))])
    dv50 = float(diams[int(np.searchsorted(cum, 0.5))])
    dv90 = float(diams[int(np.searchsorted(cum, 0.9))])
    span = float((dv90 - dv10) / dv50) if dv50 > 0 else 0.0
    return {"dv10": dv10, "dv50": dv50, "dv90": dv90, "span": span, "count": len(diams)}


def _build_excel(req: "SaveProjectRequest", excel_path: str, lang='th'):
    """
    Build Excel report with Thai language and side-by-side layout:
    - Left: Dashboard (Summary stats + chart)
    - Right: Slide data tables
    """
    # Initialize language config
    L = LanguageConfig(lang)
    
    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()

    # Compute stats for all slides
    stats_list = [_compute_slide_stats(s.droplets) for s in req.slides]

    # ── Sheet 1: Summary Dashboard ───────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    # Title (top-left)
    ws.merge_cells("A1:F1")
    ws["A1"] = f"DropDetect AI - {req.project_name}"
    ws["A1"].font = Font(bold=True, size=16)
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells("A2:F2")
    ws["A2"] = L.get('title')
    ws["A2"].font = Font(italic=True, size=12)
    ws.row_dimensions[2].height = 20

    # Info
    ws.merge_cells("A3:F3")
    ws["A3"] = f"{L.get('target')}: {req.target_size} µm  |  {L.get('total_slides')}: {len(req.slides)}  |  {L.get('generated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.row_dimensions[3].height = 18

    # SLIDE SUMMARY section (left side - Dashboard)
    ws.merge_cells("A5:F5")
    ws["A5"] = L.get('slide_summary')
    ws["A5"].font = Font(bold=True, size=12)
    ws.row_dimensions[5].height = 24

    # Header (left side)
    header_row = 6
    headers = [L.get('slide'), L.get('dv01'), L.get('dv05'), L.get('dv09'), L.get('span'), L.get('count'), L.get('who_status'), L.get('in_range_pct')]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    ws.row_dimensions[header_row].height = 22

    # Data rows (left side)
    for row_idx, (slide, st) in enumerate(zip(req.slides, stats_list), header_row + 1):
        who_pass = st["dv50"] is not None and 10 <= st["dv50"] <= 30
        who_text = L.get('pass') if who_pass else L.get('fail')
        in_range = len([d for d in slide.droplets if 10 <= d <= 30])
        in_range_pct = (in_range / len(slide.droplets) * 100) if slide.droplets else 0

        ws.cell(row=row_idx, column=1, value=slide.name).font = Font(bold=True, size=10)
        ws.cell(row=row_idx, column=2, value=round(st['dv10'], 3) if st['dv10'] else "N/A")
        ws.cell(row=row_idx, column=3, value=round(st['dv50'], 3) if st['dv50'] else "N/A")
        ws.cell(row=row_idx, column=4, value=round(st['dv90'], 3) if st['dv90'] else "N/A")
        ws.cell(row=row_idx, column=5, value=round(st['span'], 4) if st['span'] else "N/A")
        ws.cell(row=row_idx, column=6, value=st['count'])
        
        who_cell = ws.cell(row=row_idx, column=7, value=who_text)
        who_cell.font = Font(bold=True, color="008000" if who_pass else "FF0000")
        
        pct_cell = ws.cell(row=row_idx, column=8, value=round(in_range_pct, 1))
        pct_cell.number_format = "0.0%"
        
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws.row_dimensions[row_idx].height = 20

    # Column widths (left side)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 12

    # OVERALL STATISTICS (left side, below table)
    overall_row = header_row + len(req.slides) + 2
    total_droplets = sum(st['count'] for st in stats_list)
    avg_vmd = sum(st['dv50'] for st in stats_list if st['dv50']) / len([st for st in stats_list if st['dv50']]) if any(st['dv50'] for st in stats_list) else 0
    avg_span = sum(st['span'] for st in stats_list if st['span']) / len([st for st in stats_list if st['span']]) if any(st['span'] for st in stats_list) else 0
    who_pass_count = sum(1 for st in stats_list if st['dv50'] and 10 <= st['dv50'] <= 30)

    ws.merge_cells(f"A{overall_row}:F{overall_row}")
    ws[f"A{overall_row}"] = L.get('overall_stats')
    ws[f"A{overall_row}"].font = Font(bold=True, size=12)
    ws.row_dimensions[overall_row].height = 24

    overall_data = [
        (L.get('total_droplets'), total_droplets),
        (L.get('avg_vmd'), round(avg_vmd, 3)),
        (L.get('avg_span'), round(avg_span, 4)),
        (L.get('who_compliant'), f"{who_pass_count}/{len(req.slides)}"),
    ]

    for col_idx, (label, value) in enumerate(overall_data, 1):
        row = overall_row + 1
        ws.cell(row=row, column=col_idx*2-1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=col_idx*2, value=value).font = Font(size=10)

    # Add VMD Distribution Chart (left side, below stats)
    chart_row_start = overall_row + 3
    ws.merge_cells(f"A{chart_row_start}:F{chart_row_start}")
    ws[f"A{chart_row_start}"] = L.get('vmd_chart_title')
    ws[f"A{chart_row_start}"].font = Font(bold=True, size=12)
    ws.row_dimensions[chart_row_start].height = 24

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = L.get('vmd_chart_title')
    chart.y_axis.title = L.get('vmd_chart_y')
    chart.x_axis.title = L.get('vmd_chart_x')

    data_ref = Reference(ws, min_col=3, min_row=header_row+1, max_row=header_row+len(req.slides), max_col=3)
    cat_ref = Reference(ws, min_col=1, min_row=header_row+1, max_row=header_row+len(req.slides))

    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cat_ref)
    chart.shape = 4
    chart.width = 14
    chart.height = 10

    ws.add_chart(chart, f"A{chart_row_start+1}")

    # ── Per-slide sheets (right side layout) ──────────────────────────────────
    for slide_idx, slide in enumerate(req.slides):
        if not slide.droplets:
            continue

        diams = sorted(slide.droplets)  # Sort from small to large
        st = stats_list[slide_idx]
        sheet_name = slide.name[:31]
        ws2 = wb.create_sheet(title=sheet_name)

        # Title
        ws2.merge_cells("A1:F1")
        ws2["A1"] = f"{slide.name} - {L.get('droplet_data')}"
        ws2["A1"].font = Font(bold=True, size=14)
        ws2.row_dimensions[1].height = 28

        # Stats summary (left side)
        stat_labels = [
            (L.get('dv01'), round(st['dv10'], 3) if st['dv10'] else "N/A"),
            (L.get('dv05'), round(st['dv50'], 3) if st['dv50'] else "N/A"),
            (L.get('dv09'), round(st['dv90'], 3) if st['dv90'] else "N/A"),
            (L.get('span'), round(st['span'], 4) if st['span'] else "N/A"),
            (L.get('count'), st["count"]),
        ]
        for r, (lbl, val) in enumerate(stat_labels, 2):
            ws2.cell(row=r, column=1, value=lbl).font = Font(bold=True, size=10)
            ws2.cell(row=r, column=2, value=val).font = Font(size=10)
            ws2.row_dimensions[r].height = 18

        # WHO status
        who_pass = st['dv50'] and 10 <= st['dv50'] <= 30
        in_range = len([d for d in diams if 10 <= d <= 30])
        in_range_pct = (in_range / len(diams) * 100) if diams else 0
        
        ws2.cell(row=2, column=3, value=L.get('who_status_short')).font = Font(bold=True, size=10)
        ws2.cell(row=2, column=4, value=L.get('pass') if who_pass else L.get('fail')).font = Font(bold=True, color="008000" if who_pass else "FF0000")
        
        ws2.cell(row=3, column=3, value=L.get('in_range')).font = Font(bold=True, size=10)
        ws2.cell(row=3, column=4, value=round(in_range_pct, 1)).number_format = "0.0%"

        # Raw data header
        hdr_row = 5
        headers = [L.get('hash'), L.get('diameter'), L.get('who_range'), L.get('volume'), L.get('spread_factor'), L.get('cumulative_pct')]
        for ci, txt in enumerate(headers, 1):
            c = ws2.cell(row=hdr_row, column=ci, value=txt)
            c.font = Font(bold=True, size=10)
            c.alignment = Alignment(horizontal="center")
            c.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws2.row_dimensions[hdr_row].height = 20

        # Data rows (sorted small to large)
        cum_count = 0
        for i, d in enumerate(diams, 1):
            cum_count += 1
            vol = (np.pi / 6) * (d ** 3)
            sf = 0.86 if d > 20 else 0.80 if d >= 15 else 0.75 if d >= 10 else 0.70
            in_range_flag = 10.0 <= d <= 30.0
            cum_pct = (cum_count / len(diams) * 100)
            row_num = hdr_row + i
            
            ws2.cell(row=row_num, column=1, value=i)
            ws2.cell(row=row_num, column=2, value=round(d, 3)).number_format = "0.000"
            ws2.cell(row=row_num, column=3, value=L.get('yes') if in_range_flag else L.get('no')).font = Font(color="008000" if in_range_flag else "FF0000")
            ws2.cell(row=row_num, column=4, value=round(vol, 1)).number_format = "0.0"
            ws2.cell(row=row_num, column=5, value=sf).number_format = "0.00"
            ws2.cell(row=row_num, column=6, value=round(cum_pct, 1)).number_format = "0.0%"
            
            for ci in range(1, 7):
                cell = ws2.cell(row=row_num, column=ci)
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            ws2.row_dimensions[row_num].height = 16

        # Column widths
        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 14
        ws2.column_dimensions["C"].width = 12
        ws2.column_dimensions["D"].width = 14
        ws2.column_dimensions["E"].width = 14
        ws2.column_dimensions["F"].width = 12

        # Add distribution chart
        bins = {"<10": 0, "10-15": 0, "15-20": 0, "20-25": 0, "25-30": 0, ">30": 0}
        for d in diams:
            if d < 10: bins["<10"] += 1
            elif d < 15: bins["10-15"] += 1
            elif d < 20: bins["15-20"] += 1
            elif d < 25: bins["20-25"] += 1
            elif d < 30: bins["25-30"] += 1
            else: bins[">30"] += 1

        start_col = 8
        start_row = hdr_row + len(diams) + 2
        ws2.cell(row=start_row, column=start_col, value=L.get('size_range')).font = Font(bold=True)
        ws2.cell(row=start_row, column=start_col+1, value=L.get('count_label')).font = Font(bold=True)

        for i, (label, count) in enumerate(zip(list(bins.keys()), list(bins.values())), start_row+1):
            ws2.cell(row=i, column=start_col, value=label)
            ws2.cell(row=i, column=start_col+1, value=count)

        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = L.get('size_distribution')
        chart2.y_axis.title = L.get('count_label')
        chart2.x_axis.title = L.get('size_range')

        data_ref2 = Reference(ws2, min_col=start_col+1, min_row=start_row, max_row=start_row+len(bins), max_col=start_col+1)
        cat_ref2 = Reference(ws2, min_col=start_col, min_row=start_row+1, max_row=start_row+len(bins))

        chart2.add_data(data_ref2, titles_from_data=False)
        chart2.set_categories(cat_ref2)
        chart2.shape = 4
        chart2.width = 12
        chart2.height = 8

        ws2.add_chart(chart2, f"A{start_row+1}")

    # Save
    os.makedirs(os.path.dirname(excel_path) if os.path.dirname(excel_path) else ".", exist_ok=True)
    wb.save(excel_path)


def _add_summary_chart(wb, ws, slides, stats_list, header_row, num_slides):
    """Add distribution chart to Summary sheet."""

    # Create chart data range
    chart_row_start = header_row + num_slides + 4
    chart_row_end = header_row + num_slides + 4 + len(slides) - 1
    
    # Add chart title row
    ws.merge_cells(f"A{chart_row_start}:H{chart_row_start}")
    ws[f"A{chart_row_start}"] = "VMD Distribution by Slide"
    ws[f"A{chart_row_start}"].font = Font(bold=True, size=12)
    ws.row_dimensions[chart_row_start].height = 24
    
    # Create chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "VMD Comparison (µm)"
    chart.y_axis.title = "VMD (µm)"
    chart.x_axis.title = "Slide"
    
    # Data range (VMD values)
    data_ref = Reference(ws, min_col=3, min_row=header_row+1, max_row=header_row+num_slides, max_col=3)
    cat_ref = Reference(ws, min_col=1, min_row=header_row+1, max_row=header_row+num_slides)
    
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cat_ref)
    chart.shape = 4
    chart.width = 12
    chart.height = 8
    
    # Place chart
    ws.add_chart(chart, f"A{chart_row_start+1}")


def _add_slide_chart(wb, ws, diams, hdr_row, num_rows):
    """Add distribution histogram to slide sheet."""

    # Calculate distribution bins
    bins = {"<10": 0, "10-15": 0, "15-20": 0, "20-25": 0, "25-30": 0, ">30": 0}
    for d in diams:
        if d < 10:
            bins["<10"] += 1
        elif d < 15:
            bins["10-15"] += 1
        elif d < 20:
            bins["15-20"] += 1
        elif d < 25:
            bins["20-25"] += 1
        elif d < 30:
            bins["25-30"] += 1
        else:
            bins[">30"] += 1
    
    # Add bin data to sheet (hidden columns)
    start_col = 8  # Column H
    start_row = hdr_row + num_rows + 2
    
    ws.cell(row=start_row, column=start_col, value="Size Range (µm)")
    ws.cell(row=start_row, column=start_col, value="Size Range (µm)").font = Font(bold=True)
    ws.cell(row=start_row, column=start_col+1, value="Count")
    ws.cell(row=start_row, column=start_col+1, value="Count").font = Font(bold=True)
    
    bin_labels = list(bins.keys())
    bin_values = list(bins.values())
    
    for i, (label, count) in enumerate(zip(bin_labels, bin_values), start_row+1):
        ws.cell(row=i, column=start_col, value=label)
        ws.cell(row=i, column=start_col+1, value=count)
    
    # Create chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Droplet Size Distribution"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Size Range (µm)"
    
    # Data range
    data_ref = Reference(ws, min_col=start_col+1, min_row=start_row, max_row=start_row+len(bins), max_col=start_col+1)
    cat_ref = Reference(ws, min_col=start_col, min_row=start_row+1, max_row=start_row+len(bins))
    
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cat_ref)
    chart.shape = 4
    chart.width = 12
    chart.height = 8
    
    # Place chart
    ws.add_chart(chart, f"A{start_row+1}")


@app.post("/api/cleanup-autosave")
async def cleanup_autosave(project_name: str):
    """Clean up auto-save files for a specific project after successful manual save."""
    import shutil
    from pathlib import Path
    
    try:
        docs = Path.home() / "Documents" / "DropDetect_Projects"
        autosave_dir = docs / ".autosave"
        
        if autosave_dir.exists():
            # Delete autosave files older than 7 days
            now = datetime.now()
            for f in autosave_dir.glob(f"{project_name}*"):
                try:
                    # Delete files older than 7 days
                    mtime = datetime.fromtimestamp(f.stat().st_mtime)
                    age = now - mtime
                    if age.days > 7:
                        f.unlink()
                        logger.info(f"Cleaned up old autosave: {f.name}")
                except Exception as e:
                    logger.error(f"Failed to cleanup {f.name}: {e}")
        
        return {"status": "success"}
    except Exception as e:
        logger.error(f"Cleanup error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/save-project")
async def save_project(req: SaveProjectRequest):
    import shutil
    from pathlib import Path
    
    project_base = os.path.join(req.save_directory, req.project_name)
    excel_path   = f"{project_base}_Report.xlsx"
    drop_path    = f"{project_base}.drop"

    os.makedirs(req.save_directory, exist_ok=True)
    
    try:
        # Build Excel with language setting (default Thai)
        lang = req.language if hasattr(req, 'language') else 'th'
        _build_excel(req, excel_path, lang=lang)
    except PermissionError as e:
        logger.error(f"Excel file is locked: {e}")
        raise HTTPException(status_code=409, detail="Excel file is currently in use. Please close it and try again.")
    except Exception as e:
        logger.error(f"Excel build failed: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to build Excel: {str(e)}")

    try:
        # Create .drop file (ZIP) with file share mode
        with zipfile.ZipFile(drop_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            zipf.writestr("project.json", req.model_dump_json(indent=4))
            # Add Excel with error handling for file locks
            try:
                zipf.write(excel_path, arcname=os.path.basename(excel_path))
            except PermissionError:
                logger.warning("Could not add Excel to ZIP (file locked), continuing with JSON only")
    except Exception as e:
        logger.error(f"Failed to create .drop file: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to create project file: {str(e)}")

    return {"status": "success", "excel_file": excel_path, "drop_file": drop_path}

@app.get("/api/load-project")
async def load_project(path: str):
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    try:
        with zipfile.ZipFile(path, 'r') as zipf:
            with zipf.open("project.json") as f:
                return json.loads(f.read().decode('utf-8'))
    except Exception as e:
        logger.error("Failed to load project %s: %s", path, e)
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/session-data")
async def get_session_data():
    # Convert numpy types to native python floats for JSON serialization
    return {"data": [float(v) for v in system.droplet_data.values()]}

@app.post("/api/update-manual-data")
async def update_manual_data(req: ManualDataRequest):
    system.update_manual_data(req.data)
    return {"status": "success", "count": len(req.data)}

@app.post("/api/reset-stats")
async def reset_stats():
    system.reset_stats()
    return {"status": "success"}

@app.websocket("/ws/stream")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()

    # Announce hardware info immediately on connect
    ram_gb = psutil.virtual_memory().total / (1024 ** 3)
    cpu_cores = psutil.cpu_count(logical=False) or 1
    await websocket.send_json({
        "type": "hardware_info",
        "profile": system.profile,
        "ram_gb": round(ram_gb, 1),
        "cpu_cores": int(cpu_cores),
        "inference_skip": system.inference_skip
    })

    camera_idx = 0
    cap = open_camera(camera_idx)
    box_annotator = sv.BoxAnnotator(color=sv.ColorPalette.from_hex(["#0a84ff", "#ff3b30"]), thickness=2)
    label_annotator = sv.LabelAnnotator(color=sv.ColorPalette.from_hex(["#0a84ff", "#ff3b30"]), text_padding=4, text_scale=0.5)
    loop = asyncio.get_running_loop()

    async def handle_commands():
        nonlocal camera_idx, cap
        try:
            while True:
                msg = await websocket.receive_json()
                action = msg.get("action")
                if action == "set_camera":
                    idx = int(msg.get("index", 0))
                    if idx != camera_idx:
                        camera_idx = idx
                        cap.release()
                        cap = open_camera(camera_idx)
                elif action == "set_lens":
                    system.load_resources(msg.get("lens"))
                elif action == "update_settings":
                    system.conf_threshold = float(msg.get("conf", 0.25))
                elif action == "reset_stats":
                    system.reset_stats()
                elif action == "take_snapshot":
                    system.pending_snapshot = True
                elif action == "set_ai_active":
                    system.is_ai_active = bool(msg.get("active", False))
                elif action == "set_profile":
                    profile = msg.get("profile", "auto")
                    if profile == "auto":
                        profile = detect_profile()
                    system.set_profile(profile)
                    # Notify frontend of the new active profile
                    await websocket.send_json({
                        "type": "hardware_info",
                        "profile": system.profile,
                        "ram_gb": round(ram_gb, 1),
                        "cpu_cores": int(cpu_cores),
                        "inference_skip": system.inference_skip
                    })
                elif action == "remove_droplet":
                    droplet_id = int(msg.get("id", 0))
                    system.droplet_data.pop(droplet_id, None)
                    system.counted_ids.discard(droplet_id)
                elif action == "export_excel":
                    # Quick export: save Excel report to Documents/DropDetect_Projects
                    from pathlib import Path
                    import os
                    docs = Path.home() / "Documents" / "DropDetect_Projects"
                    docs.mkdir(parents=True, exist_ok=True)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    excel_path = docs / f"DropDetect_QuickExport_{timestamp}.xlsx"
                    
                    # Create a simple slide from current session data
                    slide_data = SlideData(
                        id="quick_export",
                        name=f"Session {timestamp}",
                        droplets=[float(v) for v in system.droplet_data.values()]
                    )
                    
                    # Build minimal project for export
                    class QuickExportRequest(BaseModel):
                        project_name: str
                        target_size: int
                        save_directory: str
                        slides: list[SlideData]
                    
                    req = QuickExportRequest(
                        project_name=f"QuickExport_{timestamp}",
                        target_size=224,
                        save_directory=str(docs),
                        slides=[slide_data]
                    )
                    
                    # Reuse _build_excel function
                    _build_excel(req, str(excel_path))
                    
                    # Notify frontend
                    await websocket.send_json({
                        "action": "export_result",
                        "path": str(excel_path)
                    })
        except WebSocketDisconnect:
            pass  # Client disconnected cleanly from command handler
        except Exception as e:
            logger.error("WebSocket command handler error: %s", e)

    asyncio.create_task(handle_commands())

    try:
        while True:
            success, frame = cap.read()
            if success:
                system.frame_counter += 1
                should_infer = (system.frame_counter % system.inference_skip == 0)

                # Run AI inference — only on selected frames, always off the event loop
                if (system.is_ai_active or system.pending_snapshot) and system.session and should_infer:
                    # Letterbox: pad 640×480 → 640×640 (80px top+bottom, no horizontal pad)
                    pad = np.zeros((640, 640, 3), dtype=np.uint8)
                    pad[PAD_TOP:PAD_TOP + 480, :] = frame
                    img = pad.astype(np.float32) / 255.0
                    img = np.expand_dims(np.transpose(img, (2, 0, 1)), axis=0)

                    # run_in_executor keeps the event loop unblocked during 20–80ms inference
                    _img = img  # explicit capture for lambda closure
                    outputs = await loop.run_in_executor(
                        _inference_executor,
                        lambda: system.session.run(None, {system.input_name: _img})
                    )

                    preds = np.squeeze(outputs[0]).T
                    conf_mask = preds[:, 4] > system.conf_threshold
                    valid = preds[conf_mask]
                    if len(valid) > 0:
                        # scale=1.0 (no resize), undo only PAD_TOP offset
                        xyxy = np.zeros((len(valid), 4), dtype=np.float32)
                        xyxy[:, 0] = valid[:, 0] - valid[:, 2] / 2
                        xyxy[:, 1] = valid[:, 1] - valid[:, 3] / 2 - PAD_TOP
                        xyxy[:, 2] = valid[:, 0] + valid[:, 2] / 2
                        xyxy[:, 3] = valid[:, 1] + valid[:, 3] / 2 - PAD_TOP
                        detections = sv.Detections(xyxy=xyxy, confidence=valid[:, 4], class_id=np.zeros(len(valid), dtype=int))
                        detections = detections.with_nms(threshold=0.5)
                        detections = system.tracker.update_with_detections(detections)
                        if detections.tracker_id is not None:
                            labels = []
                            class_ids = []
                            for i, tid in enumerate(detections.tracker_id):
                                box = detections.xyxy[i]
                                px_avg = ((box[2]-box[0]) + (box[3]-box[1])) / 2
                                crater_um = px_avg * system.calibration_value * 1e6
                                sf = 0.86 if crater_um > 20 else 0.80 if crater_um >= 15 else 0.75 if crater_um >= 10 else 0.70
                                true_diam = crater_um * sf
                                is_standard = 10.0 <= true_diam <= 30.0
                                cid = 0 if is_standard else 1
                                class_ids.append(cid)
                                if tid not in system.counted_ids:
                                    system.droplet_data[tid] = true_diam
                                    system.counted_ids.add(tid)
                                labels.append(f"#{tid} {true_diam:.1f}um")
                            detections.class_id = np.array(class_ids)
                            frame = box_annotator.annotate(scene=frame, detections=detections)
                            frame = label_annotator.annotate(scene=frame, detections=detections, labels=labels)

                # Prepare unified droplet list for the frontend table
                unified_list = []
                for tid, diam in system.droplet_data.items():
                    unified_list.append({
                        "id": int(tid),
                        "diameter": float(diam),
                        "source": "Manual" if tid < 0 else "AI"
                    })
                # Sort by absolute ID to maintain a consistent order (Manual IDs are negative)
                unified_list.sort(key=lambda x: abs(x["id"]))

                vmd, span, count, _, out_pct = system.calculate_stats()
                _, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
                payload = {
                    "image": base64.b64encode(buffer).decode('utf-8'),
                    "vmd": float(vmd), "span": float(span), "count": int(count),
                    "out_of_bounds": float(out_pct), "ram": str(psutil.virtual_memory().percent) + "%",
                    "current_idx": int(camera_idx),
                    "session_droplets": unified_list
                }
                if system.pending_snapshot:
                    payload["is_snapshot"] = True
                    system.pending_snapshot = False
                await websocket.send_json(payload)
            await asyncio.sleep(0.01)
    except WebSocketDisconnect:
        cap.release()
    finally:
        cap.release()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
