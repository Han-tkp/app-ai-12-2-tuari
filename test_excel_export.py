"""
Test Excel Export Format - Default Style with Charts
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import numpy as np
from datetime import datetime

# Create workbook
wb = Workbook()

# Sample data
slides = [
    {"name": "Slide 1", "droplets": [10.2, 12.5, 15.3, 18.7, 22.1, 25.4, 28.9]},
    {"name": "Slide 2", "droplets": [11.1, 13.2, 16.8, 19.5, 23.7, 26.2, 29.1]},
    {"name": "Slide 3", "droplets": [9.8, 14.1, 17.2, 20.8, 24.5, 27.3, 30.5]},
]

def compute_stats(droplets):
    if not droplets:
        return {"dv10": None, "dv50": None, "dv90": None, "span": None, "count": 0}
    diams = sorted(droplets)
    vols = [(np.pi / 6) * (d ** 3) for d in diams]
    total_vol = sum(vols)
    cum = np.cumsum(vols) / total_vol
    dv10 = float(diams[int(np.searchsorted(cum, 0.1))])
    dv50 = float(diams[int(np.searchsorted(cum, 0.5))])
    dv90 = float(diams[int(np.searchsorted(cum, 0.9))])
    span = float((dv90 - dv10) / dv50) if dv50 > 0 else 0.0
    return {"dv10": dv10, "dv50": dv50, "dv90": dv90, "span": span, "count": len(diams)}

stats_list = [compute_stats(s["droplets"]) for s in slides]

# ── Sheet 1: Summary Dashboard ───────────────────────────────────────────────
ws = wb.active
ws.title = "Summary"

# Title
ws.merge_cells("A1:H1")
ws["A1"] = f"DropDetect AI - Test Project"
ws["A1"].font = Font(bold=True, size=16)
ws.row_dimensions[1].height = 30

# Subtitle
ws.merge_cells("A2:H2")
ws["A2"] = "WHO Chemical Spray Droplet Analysis Report"
ws["A2"].font = Font(italic=True, size=12)
ws.row_dimensions[2].height = 20

# Info
ws.merge_cells("A3:H3")
ws["A3"] = f"Target Size: 224 µm  |  Total Slides: 3  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
ws.row_dimensions[3].height = 18

# SLIDE SUMMARY section
ws.merge_cells("A5:H5")
ws["A5"] = "SLIDE SUMMARY"
ws["A5"].font = Font(bold=True, size=12)
ws.row_dimensions[5].height = 24

# Header
header_row = 6
headers = ["Slide", "Dv0.1 (µm)", "Dv0.5 / VMD (µm)", "Dv0.9 (µm)", "SPAN", "Count (n)", "WHO Status", "In Range %"]
for col_idx, h in enumerate(headers, 1):
    c = ws.cell(row=header_row, column=col_idx, value=h)
    c.font = Font(bold=True, size=10)
    c.alignment = Alignment(horizontal="center")
    c.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
ws.row_dimensions[header_row].height = 22

# Data rows
for row_idx, (slide, st) in enumerate(zip(slides, stats_list), header_row + 1):
    who_pass = st["dv50"] is not None and 10 <= st["dv50"] <= 30
    who_text = "Pass" if who_pass else "Fail"
    in_range = len([d for d in slide["droplets"] if 10 <= d <= 30])
    in_range_pct = (in_range / len(slide["droplets"]) * 100)
    
    ws.cell(row=row_idx, column=1, value=slide["name"]).font = Font(bold=True, size=10)
    ws.cell(row=row_idx, column=2, value=round(st['dv10'], 3))
    ws.cell(row=row_idx, column=3, value=round(st['dv50'], 3))
    ws.cell(row=row_idx, column=4, value=round(st['dv90'], 3))
    ws.cell(row=row_idx, column=5, value=round(st['span'], 4))
    ws.cell(row=row_idx, column=6, value=st['count'])
    
    who_cell = ws.cell(row=row_idx, column=7, value=who_text)
    who_cell.font = Font(bold=True, color="008000" if who_pass else "FF0000")
    pct_cell = ws.cell(row=row_idx, column=8, value=round(in_range_pct, 1))
    pct_cell.number_format = "0.0%"
    
    for col in range(1, 9):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    ws.row_dimensions[row_idx].height = 20

# Column widths
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 10
ws.column_dimensions["F"].width = 10
ws.column_dimensions["G"].width = 12
ws.column_dimensions["H"].width = 12

# OVERALL STATISTICS
overall_row = header_row + len(slides) + 2
total_droplets = sum(st['count'] for st in stats_list)
avg_vmd = sum(st['dv50'] for st in stats_list) / len(stats_list)
avg_span = sum(st['span'] for st in stats_list) / len(stats_list)
who_pass_count = sum(1 for st in stats_list if st['dv50'] and 10 <= st['dv50'] <= 30)

ws.merge_cells(f"A{overall_row}:H{overall_row}")
ws[f"A{overall_row}"] = "OVERALL STATISTICS"
ws[f"A{overall_row}"].font = Font(bold=True, size=12)
ws.row_dimensions[overall_row].height = 24

overall_data = [
    ("Total Droplets", total_droplets),
    ("Avg VMD (µm)", round(avg_vmd, 3)),
    ("Avg SPAN", round(avg_span, 4)),
    ("WHO Compliant Slides", f"{who_pass_count}/{len(slides)}"),
]
for col_idx, (label, value) in enumerate(overall_data, 1):
    row = overall_row + 1
    ws.cell(row=row, column=col_idx*2-1, value=label).font = Font(bold=True, size=10)
    ws.cell(row=row, column=col_idx*2, value=value).font = Font(size=10)

# Add VMD Distribution Chart
chart_row_start = header_row + len(slides) + 4
ws.merge_cells(f"A{chart_row_start}:H{chart_row_start}")
ws[f"A{chart_row_start}"] = "VMD Distribution by Slide"
ws[f"A{chart_row_start}"].font = Font(bold=True, size=12)
ws.row_dimensions[chart_row_start].height = 24

chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "VMD Comparison (µm)"
chart.y_axis.title = "VMD (µm)"
chart.x_axis.title = "Slide"
data_ref = Reference(ws, min_col=3, min_row=header_row+1, max_row=header_row+len(slides), max_col=3)
cat_ref = Reference(ws, min_col=1, min_row=header_row+1, max_row=header_row+len(slides))
chart.add_data(data_ref, titles_from_data=False)
chart.set_categories(cat_ref)
chart.shape = 4
chart.width = 12
chart.height = 8
ws.add_chart(chart, f"A{chart_row_start+1}")

# ── Per-slide sheets ─────────────────────────────────────────────────────────
for slide_idx, slide in enumerate(slides):
    diams = sorted(slide["droplets"])
    st = stats_list[slide_idx]
    sheet_name = slide["name"][:31]
    ws2 = wb.create_sheet(title=sheet_name)
    
    # Title
    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"{slide['name']} - Droplet Data (Sorted by Size)"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.row_dimensions[1].height = 28
    
    # Stats summary
    stat_labels = [
        ("Dv0.1 (µm)", round(st['dv10'], 3)),
        ("Dv0.5 / VMD (µm)", round(st['dv50'], 3)),
        ("Dv0.9 (µm)", round(st['dv90'], 3)),
        ("SPAN", round(st['span'], 4)),
        ("Count (n)", st["count"]),
    ]
    for r, (lbl, val) in enumerate(stat_labels, 2):
        ws2.cell(row=r, column=1, value=lbl).font = Font(bold=True, size=10)
        ws2.cell(row=r, column=2, value=val).font = Font(size=10)
        ws2.row_dimensions[r].height = 18
    
    # WHO status
    who_pass = st['dv50'] and 10 <= st['dv50'] <= 30
    in_range = len([d for d in diams if 10 <= d <= 30])
    in_range_pct = (in_range / len(diams) * 100)
    ws2.cell(row=2, column=3, value="WHO Status").font = Font(bold=True, size=10)
    ws2.cell(row=2, column=4, value="Pass" if who_pass else "Fail").font = Font(bold=True, color="008000" if who_pass else "FF0000")
    ws2.cell(row=3, column=3, value="In Range").font = Font(bold=True, size=10)
    ws2.cell(row=3, column=4, value=round(in_range_pct, 1)).number_format = "0.0%"
    
    # Raw data header
    hdr_row = 5
    headers = ["#", "Diameter (µm)", "WHO Range", "Volume (µm³)", "Spread Factor", "Cumulative %"]
    for ci, txt in enumerate(headers, 1):
        c = ws2.cell(row=hdr_row, column=ci, value=txt)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    ws2.row_dimensions[hdr_row].height = 20
    
    # Data rows
    cum_count = 0
    for i, d in enumerate(diams, 1):
        cum_count += 1
        vol = (np.pi / 6) * (d ** 3)
        sf = 0.86 if d > 20 else 0.80 if d >= 15 else 0.75 if d >= 10 else 0.70
        in_range_flag = 10.0 <= d <= 30.0
        cum_pct = (cum_count / len(diams) * 100)
        row_num = hdr_row + i
        
        ws2.cell(row=row_num, column=1, value=i)
        ws2.cell(row=row_num, column=2, value=round(d, 3)).number_format = "0.000"
        ws2.cell(row=row_num, column=3, value="Yes" if in_range_flag else "No").font = Font(color="008000" if in_range_flag else "FF0000")
        ws2.cell(row=row_num, column=4, value=round(vol, 1)).number_format = "0.0"
        ws2.cell(row=row_num, column=5, value=sf).number_format = "0.00"
        ws2.cell(row=row_num, column=6, value=round(cum_pct, 1)).number_format = "0.0%"
        
        for ci in range(1, 7):
            cell = ws2.cell(row=row_num, column=ci)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws2.row_dimensions[row_num].height = 16
    
    # Column widths
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 12
    
    # Add distribution chart
    bins = {"<10": 0, "10-15": 0, "15-20": 0, "20-25": 0, "25-30": 0, ">30": 0}
    for d in diams:
        if d < 10: bins["<10"] += 1
        elif d < 15: bins["10-15"] += 1
        elif d < 20: bins["15-20"] += 1
        elif d < 25: bins["20-25"] += 1
        elif d < 30: bins["25-30"] += 1
        else: bins[">30"] += 1
    
    start_col = 8
    start_row = hdr_row + len(diams) + 2
    ws2.cell(row=start_row, column=start_col, value="Size Range (µm)").font = Font(bold=True)
    ws2.cell(row=start_row, column=start_col+1, value="Count").font = Font(bold=True)
    
    for i, (label, count) in enumerate(zip(list(bins.keys()), list(bins.values())), start_row+1):
        ws2.cell(row=i, column=start_col, value=label)
        ws2.cell(row=i, column=start_col+1, value=count)
    
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "Droplet Size Distribution"
    chart2.y_axis.title = "Count"
    chart2.x_axis.title = "Size Range (µm)"
    data_ref2 = Reference(ws2, min_col=start_col+1, min_row=start_row, max_row=start_row+len(bins), max_col=start_col+1)
    cat_ref2 = Reference(ws2, min_col=start_col, min_row=start_row+1, max_row=start_row+len(bins))
    chart2.add_data(data_ref2, titles_from_data=False)
    chart2.set_categories(cat_ref2)
    chart2.shape = 4
    chart2.width = 12
    chart2.height = 8
    ws2.add_chart(chart2, f"A{start_row+1}")

# Save
wb.save("C:/Users/h4n/Desktop/webappsdesktop/datafordevapp/excel/TestExport_Default_WithCharts.xlsx")
print("Excel file created: TestExport_Default_WithCharts.xlsx")
print("\nStructure:")
print("  Sheet 1: Summary (Dashboard + VMD Chart)")
print("  Sheet 2: Slide 1 (sorted data + Distribution Chart)")
print("  Sheet 3: Slide 2 (sorted data + Distribution Chart)")
print("  Sheet 4: Slide 3 (sorted data + Distribution Chart)")
